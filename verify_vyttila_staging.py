import openpyxl
import psycopg2

conn = psycopg2.connect(
    host='aws-1-ap-south-1.pooler.supabase.com',
    port=6543,
    user='postgres.zhatpmhkegzyrrmtdzhm',
    password='Defab_staging_123',
    dbname='postgres',
    sslmode='require'
)
cur = conn.cursor()

# Get Vyttila warehouse ID
cur.execute("SELECT w.id, w.name FROM warehouses w WHERE w.name ILIKE '%vyttila%'")
warehouses = cur.fetchall()
print('Vyttila warehouses:', warehouses)

if not warehouses:
    print('ERROR: No Vyttila warehouse found — has the import been run on this DB yet?')
    conn.close()
    exit(1)

wh_id = warehouses[0][0]
print('Using warehouse:', warehouses[0])
print()

# Get all stock for this warehouse from DB
cur.execute("""
    SELECT
        v.variant_code,
        v.price,
        v.cost_price,
        p.name as product_name,
        c.name as category_name,
        s.quantity,
        s.id as stock_id,
        v.id as variant_id
    FROM stocks s
    JOIN variants v ON v.id = s.variant_id
    JOIN products p ON p.id = v.product_id
    JOIN categories c ON c.id = p.category_id
    WHERE s.warehouse_id = %s
    ORDER BY c.name, v.variant_code
""", (wh_id,))
db_rows = cur.fetchall()
print('Total DB stock rows for Vyttila:', len(db_rows))

# Build DB lookup: (variant_code, category) -> dict
# Use (code, lower_category) as key to handle codes appearing in multiple sheets
db_map_by_cat = {}  # (code, cat) -> dict
db_map_by_code = {} # code -> list (for simple single-code lookup)
for row in db_rows:
    code, price, cp, prod, cat, qty, stock_id, vid = row
    key = (code, cat.lower().strip())
    db_map_by_cat[key] = {
        'price': float(price) if price else 0,
        'cp': float(cp) if cp else 0,
        'qty': float(qty) if qty else 0,
        'product': prod,
        'stock_id': stock_id,
        'variant_id': vid,
    }
    if code not in db_map_by_code:
        db_map_by_code[code] = []
    db_map_by_code[code].append(db_map_by_cat[key])

# Parse Excel
excel_path = r'd:\QMark\defab_erp_backend\internal\migration\Defab Vyttila\LATEST_STOCK_MAY2026_WITH_GST.xlsx'
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

excel_map_by_cat = {}  # (code, sheet_name_lower) -> dict
excel_total_qty = 0

for sname in wb.sheetnames:
    ws = wb[sname]
    all_rows = list(ws.iter_rows(values_only=True))
    header_row = -1
    for i, row in enumerate(all_rows[:3]):
        if len(row) > 1 and row[1] is not None and str(row[1]).strip().upper() == 'CODE':
            header_row = i
            break
    if header_row < 0:
        continue
    sheet_count = 0
    seen_in_sheet = {}
    for i, row in enumerate(all_rows):
        if i <= header_row or len(row) < 8:
            continue
        code = row[1]
        if code is None:
            continue
        try:
            code_int = int(float(str(code)))
        except:
            continue
        if code_int == 0:
            continue
        sp_excl = float(row[3]) if row[3] else 0
        cp = float(row[6]) if row[6] else 0
        qty = float(row[7]) if row[7] else 0
        if cp <= 0:
            cp = sp_excl
        # Aggregate duplicates within same sheet
        key = (code_int, sname.lower().strip())
        if key in seen_in_sheet:
            seen_in_sheet[key]['qty'] += qty
        else:
            seen_in_sheet[key] = {
                'price': round(sp_excl, 4),
                'cp': round(cp, 4),
                'qty': qty,
                'sheet': sname,
            }
        excel_total_qty += qty
        sheet_count += 1
    excel_map_by_cat.update(seen_in_sheet)
    print(f'  Sheet: {sname} -> {sheet_count} rows')

wb.close()

print(f'\nExcel unique (code, sheet) pairs: {len(excel_map_by_cat)}')
print(f'Excel total qty: {excel_total_qty}')
print()

# Compare using (code, category) keys
price_mismatches = []
cp_mismatches = []
qty_mismatches = []
missing_in_db = []
extra_in_db = []

for key, ex in excel_map_by_cat.items():
    code, sheet_lower = key
    if key not in db_map_by_cat:
        missing_in_db.append((code, ex['sheet']))
    else:
        db = db_map_by_cat[key]
        if abs(ex['price'] - db['price']) > 0.05:
            price_mismatches.append((code, ex['sheet'], ex['price'], db['price']))
        if abs(ex['cp'] - db['cp']) > 0.05:
            cp_mismatches.append((code, ex['sheet'], ex['cp'], db['cp']))
        if abs(ex['qty'] - db['qty']) > 0.01:
            qty_mismatches.append((code, ex['sheet'], ex['qty'], db['qty']))

for key in db_map_by_cat:
    if key not in excel_map_by_cat:
        code, cat = key
        extra_in_db.append((code, cat))

print('=' * 60)
print(f'Missing in DB (in Excel but not in DB): {len(missing_in_db)}')
print(f'Extra in DB (in DB but not in Excel):   {len(extra_in_db)}')
print(f'Price mismatches:                       {len(price_mismatches)}')
print(f'CP mismatches:                          {len(cp_mismatches)}')
print(f'QTY mismatches:                         {len(qty_mismatches)}')
print('=' * 60)

if missing_in_db:
    print('\nMissing in DB:')
    for c, sheet in missing_in_db[:20]:
        print(f'  code={c}  sheet={sheet}')

if extra_in_db:
    print('\nExtra in DB:')
    for c, cat in extra_in_db[:20]:
        print(f'  code={c}  category={cat}')

if price_mismatches:
    print('\nPrice mismatches (Excel vs DB):')
    for c, sheet, ep, dp in price_mismatches[:30]:
        print(f'  code={c}  sheet={sheet}  excel={ep}  db={dp}  diff={round(ep-dp,2)}')

if cp_mismatches:
    print('\nCP mismatches:')
    for c, sheet, ep, dp in cp_mismatches[:30]:
        print(f'  code={c}  sheet={sheet}  excel_cp={ep}  db_cp={dp}  diff={round(ep-dp,2)}')

if qty_mismatches:
    print('\nQTY mismatches:')
    for c, sheet, eq, dq in qty_mismatches[:30]:
        print(f'  code={c}  sheet={sheet}  excel_qty={eq}  db_qty={dq}  diff={round(eq-dq,2)}')

all_good = (not missing_in_db and not price_mismatches and not cp_mismatches and not qty_mismatches)
if all_good:
    print('\n✅ ALL DATA MATCHES PERFECTLY!')
else:
    print(f'\n⚠️  Issues found — see above')

cur.close()
conn.close()
