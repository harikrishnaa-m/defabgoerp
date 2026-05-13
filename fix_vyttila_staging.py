"""
Fix Vyttila variant prices, CPs, and stock quantities in the STAGING DB.
"""

import psycopg2
import openpyxl
from psycopg2.extras import execute_values

conn = psycopg2.connect(
    host='aws-1-ap-south-1.pooler.supabase.com',
    port=6543,
    user='postgres.zhatpmhkegzyrrmtdzhm',
    password='Defab_staging_123',
    dbname='postgres',
    sslmode='require'
)
conn.autocommit = False
cur = conn.cursor()

# 1. Get the Vyttila warehouse ID
cur.execute("SELECT id FROM warehouses WHERE name = 'DEFAB Vyttila'")
row = cur.fetchone()
if not row:
    print("ERROR: DEFAB Vyttila warehouse not found")
    cur.close()
    conn.close()
    exit(1)
wh_id = row[0]
print(f"Vyttila warehouse ID: {wh_id}")

# 2. Get all variants in Vyttila warehouse with category info
cur.execute("""
    SELECT v.id, v.variant_code, v.price, v.cost_price, s.quantity, s.id as stock_id, c.name as cat
    FROM stocks s
    JOIN variants v ON v.id = s.variant_id
    JOIN products p ON p.id = v.product_id
    JOIN categories c ON c.id = p.category_id
    WHERE s.warehouse_id = %s
""", (wh_id,))
db_variants = cur.fetchall()
print(f"Variants in Vyttila: {len(db_variants)}")

# Build map: (variant_code, cat_lower) -> {id, price, cp, qty, stock_id}
db_map = {}
for (vid, code, price, cp, qty, stock_id, cat) in db_variants:
    key = (code, cat.lower().strip())
    db_map[key] = {
        'id': vid,
        'price': float(price) if price else 0,
        'cp': float(cp) if cp else 0,
        'qty': float(qty) if qty else 0,
        'stock_id': stock_id
    }

# 3. Parse Excel
excel_path = r'd:\QMark\defab_erp_backend\internal\migration\Defab Vyttila\LATEST_STOCK_MAY2026_WITH_GST.xlsx'
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

excel_map = {}
for sname in wb.sheetnames:
    ws = wb[sname]
    all_rows = list(ws.iter_rows(values_only=True))
    header_row = -1
    for i, row in enumerate(all_rows[:3]):
        if len(row) > 1 and row[1] is not None and str(row[1]).strip().upper() == 'CODE':
            header_row = i
            break
    if header_row < 0:
        continue
    seen = {}
    for i, row in enumerate(all_rows):
        if i <= header_row or len(row) < 8:
            continue
        code = row[1]
        if code is None:
            continue
        try:
            code_int = int(float(str(code)))
        except:
            continue
        if code_int == 0:
            continue
        sp_excl = float(row[3]) if row[3] else 0
        cp = float(row[6]) if row[6] else 0
        qty = float(row[7]) if row[7] else 0
        if cp <= 0:
            cp = sp_excl
        key = (code_int, sname.lower().strip())
        if key in seen:
            seen[key]['qty'] += qty
        else:
            seen[key] = {'price': round(sp_excl, 4), 'cp': round(cp, 4), 'qty': qty}
    excel_map.update(seen)

wb.close()
print(f"Excel (code, sheet) pairs: {len(excel_map)}")

# 4. Compute updates
price_updates = []   # (new_price, new_cp, variant_id)
stock_updates = []   # (new_qty, stock_id, code)

price_ok = qty_ok = 0

for key, ex in excel_map.items():
    if key not in db_map:
        print(f"WARNING: {key} not found in DB — skipping")
        continue
    db = db_map[key]
    if abs(ex['price'] - db['price']) > 0.05 or abs(ex['cp'] - db['cp']) > 0.05:
        price_updates.append((ex['price'], ex['cp'], db['id']))
    else:
        price_ok += 1
    if abs(ex['qty'] - db['qty']) > 0.01:
        stock_updates.append((ex['qty'], db['stock_id'], key[0]))
    else:
        qty_ok += 1

print(f"\nPrice/CP updates needed: {len(price_updates)}")
print(f"Price/CP already correct: {price_ok}")
print(f"Stock qty updates needed: {len(stock_updates)}")
print(f"Stock qty already correct: {qty_ok}")

if not price_updates and not stock_updates:
    print("\nNothing to fix!")
    cur.close()
    conn.close()
    exit(0)

print("\nApplying updates...")

# 5. Batch UPDATE variants
if price_updates:
    execute_values(
        cur,
        """
        UPDATE variants AS v
        SET price = data.price,
            cost_price = data.cost_price,
            updated_at = NOW()
        FROM (VALUES %s) AS data(price, cost_price, id)
        WHERE v.id = data.id::uuid
        """,
        price_updates,
        template="(%s, %s, %s)"
    )
    print(f"  Variant price/CP batches applied ({len(price_updates)} total rows)")

# 6. Update stock quantities
for (new_qty, stock_id, code) in stock_updates:
    cur.execute("UPDATE stocks SET quantity = %s, updated_at = NOW() WHERE id = %s", (new_qty, stock_id))
    print(f"  Stock qty updated: code={code} -> qty={new_qty}")

conn.commit()
print("\n✅ All updates committed!")

cur.close()
conn.close()
