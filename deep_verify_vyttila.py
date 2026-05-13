"""
Deep verification: categories, products, variants, prices, CPs, quantities
against LATEST_STOCK_MAY2026_WITH_GST.xlsx — staging DB
"""

import openpyxl
import psycopg2

conn = psycopg2.connect(
    host='aws-1-ap-south-1.pooler.supabase.com',
    port=6543,
    user='postgres.zhatpmhkegzyrrmtdzhm',
    password='Defab_staging_123',
    dbname='postgres',
    sslmode='require'
)
cur = conn.cursor()

# ── Get Vyttila warehouse ──────────────────────────────────────
cur.execute("SELECT id FROM warehouses WHERE name = 'DEFAB Vyttila'")
wh_id = cur.fetchone()[0]

# ── Pull everything from DB for this warehouse ─────────────────
cur.execute("""
    SELECT
        c.name  AS cat_name,
        p.name  AS prod_name,
        v.variant_code,
        v.price,
        v.cost_price,
        s.quantity
    FROM stocks s
    JOIN variants v ON v.id = s.variant_id
    JOIN products p ON p.id = v.product_id
    JOIN categories c ON c.id = p.category_id
    WHERE s.warehouse_id = %s
""", (wh_id,))
db_rows = cur.fetchall()
cur.close()
conn.close()

# Key: (cat_lower, prod_upper, code)
db_map = {}
db_categories = set()
db_products = {}   # cat_lower -> set of prod_upper

for cat, prod, code, price, cp, qty in db_rows:
    cat_l = cat.lower().strip()
    prod_u = prod.upper().strip()
    db_categories.add(cat_l)
    db_products.setdefault(cat_l, set()).add(prod_u)
    key = (cat_l, code)
    db_map[key] = {
        'prod': prod_u,
        'price': float(price) if price else 0,
        'cp': float(cp) if cp else 0,
        'qty': float(qty) if qty else 0,
    }

print(f"DB  — categories: {len(db_categories)}, variants total: {len(db_rows)}")

# ── Parse Excel ────────────────────────────────────────────────
excel_path = r'd:\QMark\defab_erp_backend\internal\migration\Defab Vyttila\LATEST_STOCK_MAY2026_WITH_GST.xlsx'
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

xl_map = {}
xl_categories = set()
xl_products = {}    # sheet_lower -> set of item_upper
xl_total_qty = 0.0

for sname in wb.sheetnames:
    ws = wb[sname]
    all_rows = list(ws.iter_rows(values_only=True))
    header_row = -1
    for i, row in enumerate(all_rows[:3]):
        if len(row) > 1 and row[1] is not None and str(row[1]).strip().upper() == 'CODE':
            header_row = i
            break
    if header_row < 0:
        continue
    cat_l = sname.lower().strip()
    xl_categories.add(cat_l)
    for i, row in enumerate(all_rows):
        if i <= header_row or len(row) < 8:
            continue
        code_raw = row[1]
        if code_raw is None:
            continue
        try:
            code = int(float(str(code_raw)))
        except:
            continue
        if code == 0:
            continue
        sp_excl = float(row[3]) if row[3] else 0
        cp      = float(row[6]) if row[6] else 0
        qty     = float(row[7]) if row[7] else 0
        item    = str(row[11]).upper().strip() if row[11] else sname.upper()
        if cp <= 0:
            cp = sp_excl
        xl_products.setdefault(cat_l, set()).add(item)
        key = (cat_l, code)
        if key in xl_map:
            xl_map[key]['qty'] += qty
        else:
            xl_map[key] = {'prod': item, 'price': round(sp_excl,4), 'cp': round(cp,4), 'qty': qty}
        xl_total_qty += qty

wb.close()

xl_total_variants = len(xl_map)
print(f"XL  — categories: {len(xl_categories)}, variants total: {xl_total_variants}")
print()

# ── CATEGORY CHECK ─────────────────────────────────────────────
missing_cats = xl_categories - db_categories
extra_cats   = db_categories - xl_categories
print("─── CATEGORIES ───────────────────────────────────────────")
print(f"  Excel sheets : {sorted(xl_categories)}")
print(f"  DB categories: {sorted(db_categories)}")
if missing_cats:
    print(f"  ❌ Missing in DB : {missing_cats}")
else:
    print("  ✅ All categories present")
if extra_cats:
    print(f"  ⚠️  Extra in DB  : {extra_cats}")

# ── PRODUCT CHECK ──────────────────────────────────────────────
print()
print("─── PRODUCTS (per category) ───────────────────────────────")
all_prods_ok = True
for cat_l in sorted(xl_categories):
    xl_p = xl_products.get(cat_l, set())
    db_p = db_products.get(cat_l, set())
    missing_p = xl_p - db_p
    extra_p   = db_p - xl_p
    if missing_p or extra_p:
        all_prods_ok = False
        print(f"  [{cat_l}]")
        if missing_p: print(f"    ❌ Missing in DB : {missing_p}")
        if extra_p:   print(f"    ⚠️  Extra in DB  : {extra_p}")
    else:
        print(f"  ✅ [{cat_l}] — {len(xl_p)} product(s) all match")

# ── VARIANT / PRICE / CP / QTY CHECK ──────────────────────────
print()
print("─── VARIANTS / PRICES / CPs / QUANTITIES ──────────────────")
missing_variants = []
price_mismatches = []
cp_mismatches    = []
qty_mismatches   = []
extra_variants   = []

for key, ex in xl_map.items():
    if key not in db_map:
        missing_variants.append(key)
        continue
    db = db_map[key]
    if abs(ex['price'] - db['price']) > 0.05:
        price_mismatches.append((key, ex['price'], db['price']))
    if abs(ex['cp'] - db['cp']) > 0.05:
        cp_mismatches.append((key, ex['cp'], db['cp']))
    if abs(ex['qty'] - db['qty']) > 0.01:
        qty_mismatches.append((key, ex['qty'], db['qty']))

for key in db_map:
    if key not in xl_map:
        extra_variants.append(key)

print(f"  Variants in Excel : {len(xl_map)}")
print(f"  Variants in DB    : {len(db_rows)}")
print(f"  Missing in DB     : {len(missing_variants)}")
print(f"  Extra in DB       : {len(extra_variants)}")
print(f"  Price mismatches  : {len(price_mismatches)}")
print(f"  CP mismatches     : {len(cp_mismatches)}")
print(f"  QTY mismatches    : {len(qty_mismatches)}")

if missing_variants:
    print(f"\n  ❌ Missing variants (first 20):")
    for k in missing_variants[:20]:
        print(f"    {k}")

if extra_variants:
    print(f"\n  ⚠️  Extra in DB (first 10):")
    for k in extra_variants[:10]:
        print(f"    {k} -> {db_map[k]}")

if price_mismatches:
    print(f"\n  ❌ Price mismatches (first 20):")
    for (cat, code), ep, dp in price_mismatches[:20]:
        print(f"    code={code} cat={cat}  excel={ep}  db={dp}  diff={round(ep-dp,2)}")

if cp_mismatches:
    print(f"\n  ❌ CP mismatches (first 20):")
    for (cat, code), ep, dp in cp_mismatches[:20]:
        print(f"    code={code} cat={cat}  excel_cp={ep}  db_cp={dp}  diff={round(ep-dp,2)}")

if qty_mismatches:
    print(f"\n  ❌ QTY mismatches:")
    for (cat, code), eq, dq in qty_mismatches:
        print(f"    code={code} cat={cat}  excel_qty={eq}  db_qty={dq}")

# ── TOTALS ─────────────────────────────────────────────────────
print()
print("─── TOTALS ────────────────────────────────────────────────")
db_total_qty = sum(v['qty'] for v in db_map.values())
print(f"  Total qty Excel : {xl_total_qty}")
print(f"  Total qty DB    : {db_total_qty}")
qty_match = abs(xl_total_qty - db_total_qty) < 0.01
print(f"  Qty totals      : {'✅ MATCH' if qty_match else '❌ MISMATCH'}")

# ── FINAL VERDICT ──────────────────────────────────────────────
print()
issues = (missing_cats or extra_cats or not all_prods_ok or
          missing_variants or price_mismatches or cp_mismatches or qty_mismatches)
if not issues:
    print("✅ FULL VERIFICATION PASSED — DB matches Excel exactly")
else:
    print("⚠️  ISSUES FOUND — see details above")
