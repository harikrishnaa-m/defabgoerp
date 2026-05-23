"""
Export stock data from DB for each branch → one Excel per category.
Format matches the existing xlsx files in internal/migration/Defab Thrippunithura/
"""

import os
import psycopg2
import openpyxl
from openpyxl.styles import Font

DB_CONFIG = {
    "host":     "aws-1-ap-south-1.pooler.supabase.com",
    "port":     6543,
    "user":     "postgres.zhatpmhkegzyrrmtdzhm",
    "password": "Defab_staging_123",
    "dbname":   "postgres",
}

BRANCHES = [
    "DEFAB Thrippunithura",
    "DEFAB Vyttila",
]

OUTPUT_ROOT = "exports"

QUERY = """
SELECT
    c.name          AS category,
    p.name          AS item_name,
    v.variant_code  AS code,
    v.price         AS mrp,
    COALESCE(SUM(s.quantity), 0) AS qty,
    v.cost_price    AS mrp_excl_gst
FROM stocks s
JOIN variants   v  ON s.variant_id   = v.id
JOIN products   p  ON v.product_id   = p.id
JOIN categories c  ON p.category_id  = c.id
JOIN warehouses w  ON s.warehouse_id = w.id
JOIN branches   b  ON w.branch_id    = b.id
WHERE b.name = %s
GROUP BY c.name, p.name, v.variant_code, v.price, v.cost_price
ORDER BY c.name, p.name, v.variant_code
"""


def write_category_xlsx(branch_name: str, category: str, rows: list, out_dir: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    # Row 1 – title
    ws.append([f"{category} — {branch_name}"])
    ws["A1"].font = Font(bold=True)

    # Row 2 – headers
    headers = ["SL No", "Items", "Code", "MRP", "Qty", "MRP Excluding GST"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # Data rows
    for sl, (item, code, mrp, qty, cost_price) in enumerate(rows, start=1):
        ws.append([sl, item, code, float(mrp) if mrp else 0,
                   float(qty), float(cost_price) if cost_price else 0])

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 20

    safe_name = category.replace("/", "-").replace("\\", "-")
    path = os.path.join(out_dir, f"{safe_name}.xlsx")
    wb.save(path)
    print(f"  Saved: {path}  ({len(rows)} rows)")


def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    for branch in BRANCHES:
        print(f"\n=== {branch} ===")
        cur.execute(QUERY, (branch,))
        all_rows = cur.fetchall()

        if not all_rows:
            print("  No stock data found.")
            continue

        # Group by category
        from collections import defaultdict
        cat_rows = defaultdict(list)
        for category, item, code, mrp, qty, cost_price in all_rows:
            cat_rows[category].append((item, code, mrp, qty, cost_price))

        out_dir = os.path.join(OUTPUT_ROOT, branch)
        os.makedirs(out_dir, exist_ok=True)

        for category, rows in sorted(cat_rows.items()):
            write_category_xlsx(branch, category, rows, out_dir)

        print(f"  Total: {len(all_rows)} rows across {len(cat_rows)} categories")

    cur.close()
    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
